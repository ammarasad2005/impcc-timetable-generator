"""Build timetables.xlsx (template layout) from solutions.json — ALL combinations."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from metrics import section_metrics, teacher_schedule, RULES, DAYS

TIMES = ["08:30-09:10", "09:10-09:50", "09:50-10:30", "10:30-10:55", "10:55-11:35", "11:35-12:15"]

def section_title(key):
    stream, year, sec = key.rsplit("-", 2)
    return f"{stream}-{year} (Section-{sec})"

def build_sheet(ws, meta, tt):
    r = 1
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9E2F3")
    title_fill = PatternFill("solid", fgColor="2F5597")
    for sec_key in meta["section_order"]:
        grid = tt[sec_key]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c = ws.cell(row=r, column=1, value=section_title(sec_key))
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = title_fill
        c.alignment = Alignment(horizontal="center")
        r += 1
        heads = ["DAYS", "Period-1", "Period-2", "Period-3", "Break", "Period-4", "Period-5"]
        for j, h in enumerate(heads):
            c = ws.cell(row=r, column=1 + j, value=h)
            c.font = Font(bold=True); c.fill = head_fill
            c.alignment = Alignment(horizontal="center"); c.border = border
        r += 1
        for j, t in enumerate(TIMES):
            c = ws.cell(row=r, column=1 + j, value=t)
            c.font = Font(size=9, color="555555")
            c.alignment = Alignment(horizontal="center"); c.border = border
        r += 1
        for d in range(5):
            row = grid[d]
            c0 = ws.cell(row=r, column=1, value=DAYS[d])
            c0.font = Font(bold=True); c0.alignment = Alignment(horizontal="center"); c0.border = border
            for j in range(6):
                if j == 3:
                    cc = ws.cell(row=r, column=1 + j, value="")
                else:
                    slot = j if j < 3 else j - 1
                    cc = ws.cell(row=r, column=1 + j, value=row[slot][0])
                    cc.alignment = Alignment(horizontal="center", wrap_text=True)
                cc.border = border
            r += 1
            c1 = ws.cell(row=r, column=1, value=""); c1.border = border
            for j in range(6):
                if j == 3:
                    cc = ws.cell(row=r, column=1 + j, value="")
                else:
                    slot = j if j < 3 else j - 1
                    cc = ws.cell(row=r, column=1 + j, value=row[slot][1])
                    cc.font = Font(size=9, color="777777")
                    cc.alignment = Alignment(horizontal="center", wrap_text=True)
                cc.border = border
            r += 1
        r += 1
    for j, w in enumerate([8, 20, 20, 20, 14, 20, 20]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w

def build_teacher_sheet(ws, tt):
    sched = teacher_schedule(tt)
    names = sorted(sched)
    r = 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    c = ws.cell(row=1, column=1, value="Teacher Schedule (this combination)")
    c.font = Font(bold=True, size=12)
    r = 3
    for name in names:
        rows = sched[name]
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        ws.cell(row=r, column=2, value=f"{len(rows)} periods")
        ws.cell(row=r, column=3, value=RULES.get(name, "")).font = Font(italic=True, size=9, color="555555")
        r += 1
        for (d, s, sec, subj) in rows:
            ws.cell(row=r, column=2, value=f"{DAYS[d]}  P{s+1}")
            ws.cell(row=r, column=3, value=sec)
            ws.cell(row=r, column=4, value=subj)
            r += 1
        r += 1
    for j, w in enumerate([34, 12, 40, 22, 30]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w

def main():
    data = json.load(open("solutions.json"))
    meta = data["meta"]
    sols = data["solutions"]
    wb = Workbook()
    wb.remove(wb.active)

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "IMPCC Inter (1st Shift) — Generated Timetables"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"{len(sols)} valid combinations. Ranked by shuffle-preference score "
                f"(lower = better; 5/4-credit subjects are never shuffled, "
                f"3-credit rarely, 2-credit freely).")
    ws["A4"] = "Rank"; ws["B4"] = "Score"; ws["C4"] = "3cr shuffled"; ws["D4"] = "2cr shuffled"; ws["E4"] = "Sheet"
    for i, sol in enumerate(sols):
        m = section_metrics(sol["timetable"])
        ws.cell(row=5 + i, column=1, value=i + 1)
        ws.cell(row=5 + i, column=2, value=sol["score"])
        ws.cell(row=5 + i, column=3, value=m.get(3, 0))
        ws.cell(row=5 + i, column=4, value=m.get(2, 0))
        ws.cell(row=5 + i, column=5, value=f"Combo {i+1:02d}")
    for j, w in enumerate([8, 10, 14, 14, 14]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w

    # All combination sheets
    for i, sol in enumerate(sols):
        wss = wb.create_sheet(f"Combo {i+1:02d}")
        build_sheet(wss, meta, sol["timetable"])

    # Best combination's teacher schedule
    wst = wb.create_sheet("Teacher Schedule (Best)")
    build_teacher_sheet(wst, sols[0]["timetable"])

    wb.save("timetables.xlsx")
    print(f"wrote timetables.xlsx: {len(sols)} combo sheets + Summary + Teacher Schedule (Best)")

if __name__ == "__main__":
    main()
